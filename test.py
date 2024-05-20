import os
import openpyxl

def search_excel_files(folder_path, search_term):
    results = []

    # يمر عبر كل ملف Excel في المجلد
    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_path = os.path.join(folder_path, filename)

            # فتح الملف Excel
            workbook = openpyxl.load_workbook(file_path)
            
            # البحث عن الورقة المناسبة (يمكن تغييرها حسب احتياجاتك)
            sheet = workbook.active
            
            # يمر عبر الصفوف والأعمدة للبحث عن البيانات
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_index, cell_value in enumerate(row, start=1):
                    if search_term in str(cell_value):
                        # حفظ تفاصيل النتيجة
                        results.append({
                            'file_name': filename,
                            'cell_value': cell_value,
                            'cell_address': sheet.cell(row=row_index, column=col_index).coordinate
                        })
    return results

# استخدام الدالة
folder_path = 'C:/Users/bassa/OneDrive/العمل/كشوفات مكتب غليل/2024/شهر 5'
search_term = 'جدة'
results = search_excel_files(folder_path, search_term)
print(results)
