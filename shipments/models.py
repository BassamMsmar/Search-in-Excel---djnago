import os
from django.db import models
from openpyxl import load_workbook


# Create your models here.
class Shipment(models.Model):
    invoice_number= models.CharField(max_length=100) 
    receipt_number = models.CharField(max_length=100)
    driver_name = models.CharField(max_length=100)       
    fare = models.IntegerField()
    destination  = models.CharField(max_length=100) 
    company = models.CharField(max_length=100) 
    shipment_number= models.CharField(max_length=100)
    note = models.CharField(max_length=200)
    date = models.DateField(max_length=100)

class ExcelFile(models.Model):
    file = models.FileField(upload_to='excel_files/')  # حقل لتحميل الملفات
    file_name = models.CharField(max_length=100, blank=True, null=True)
    creation_time = models.DateTimeField(blank=True, null=True)
    modification_time = models.DateTimeField(blank=True, null=True)
    sheet_names = models.TextField(blank=True, null=True) 
    processing_date = models.DateTimeField(auto_now_add=True ,null=True, blank=True)
    
    


    def save(self, *args, **kwargs):
        super().save(*args, **kwargs)
        file_path = self.file.path
        wb = load_workbook(filename=file_path, read_only=True)
        properties = wb.properties


        self.file_name = os.path.basename(self.file.name)
        self.creation_time = properties.created
        self.modification_time = properties.modified
        sheet_names = wb.sheetnames
        self.sheet_names = ", ".join(sheet_names)


      

        super().save(*args, **kwargs)

    def __str__(self) -> str:
        return str(self.file_name)