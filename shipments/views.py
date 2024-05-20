from django.shortcuts import render, redirect
import os
from datetime import datetime
from openpyxl import load_workbook 

from . models import Shipment, ExcelFile
from .forms import ExcelFileForm

# Create your views here.
def shipment_list(request):
    shipments = Shipment.objects.all()
    return render(request, 'shipment/shipment_list.html', {'shipments': shipments})

def inoices_list(request):
    if request.method == 'POST':
        form = ExcelFileForm(request.POST, request.FILES)
        if form.is_valid():

            forms = form.save(commit=False)
            forms.sheet_names = "bassam"
            forms.save()
            return redirect('inoices_list')
    else:
        form = ExcelFileForm()
        files = ExcelFile.objects.all()

        context = {
            'form': form,
            'files': files,
        }
    return render(request, 'shipment/list_invoices.html', {'context': context})
        

def upload_file(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']

    #     wb = openpyxl.load_workbook(excel_file)
    #     sheet = wb.active
    #     for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    #         shipment = Shipment()
    #         shipment.invoice_number = row[0].value
    #         shipment.invoice_date = row[1].value
    #         shipment.customer_name = row[2].value
    #         shipment.customer_address = row[3].value
    #         shipment.fare = row[4].value
    #         shipment.save()
    # return render(request, 'shipment/upload_file.html')